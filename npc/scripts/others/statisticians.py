#!/usr/bin/env python
import re
import subprocess
import sys
from copy import copy
import openpyxl
from openpyxl import load_workbook

# 定义要提取的数据的正则表达式模式
patterns = {
    "仿真周期数": r"total guest cycles = ([\d,]+)",
    "指令数": r"total fetch instructions = ([\d,]+)",
    "CPI": r"CPI\(cycle per instruction\) = ([\d.]+)",
    "IPC": r"IPC\(instruction per cycle\) = ([\d.]+)",
    "取数据计数器": r"total fetch data = ([\d,]+)",
    "取数据平均周期": r"average cycle of fetch data = ([\d.]+)",
    "JMP指令计数器": r"instructions for JMP type \(total = ([\d,]+),",
    "JMP指令平均执行周期": r"instructions for JMP type \(total = [\d,]+, average exec cycle = ([\d.]+)\)",
    "BRANCH指令计数器": r"instructions for BRANCH type \(total = ([\d,]+),",
    "BRANCH指令平均执行周期": r"instructions for BRANCH type \(total = [\d,]+, average exec cycle = ([\d.]+)\)",
    "LOAD指令计数器": r"instructions for LOAD type \(total = ([\d,]+),",
    "LOAD指令平均执行周期": r"instructions for LOAD type \(total = [\d,]+, average exec cycle = ([\d.]+)\)",
    "STORE指令计数器": r"instructions for STORE type \(total = ([\d,]+),",
    "STORE指令平均执行周期": r"instructions for STORE type \(total = [\d,]+, average exec cycle = ([\d.]+)\)",
    "AL指令计数器": r"instructions for AL type \(total = ([\d,]+),",
    "AL指令平均执行周期": r"instructions for AL type \(total = [\d,]+, average exec cycle = ([\d.]+)\)",
    "CSR指令计数器": r"instructions for CSR type \(total = ([\d,]+),",
    "CSR指令平均执行周期": r"instructions for CSR type \(total = [\d,]+, average exec cycle = ([\d.]+)\)",
    "ECALL/EBREAK指令计数器": r"instructions for ECALL type \(total = ([\d,]+),",
    "ECALL/EBREAK指令平均执行周期": r"instructions for ECALL type \(total = [\d,]+, average exec cycle = ([\d.]+)\)",
    "FENCE.I指令计数器": r"instructions for FENCE_I type \(total = ([\d,]+),",
    "FENCE.I指令平均执行周期": r"instructions for FENCE_I type \(total = [\d,]+, average exec cycle = ([\d.]+)\)",
    "缓存访问次数": r"total cache access = ([\d,]+)",
    "缓存命中次数": r"total cache hit = ([\d,]+)",
    "缓存命中率": r"cache hit ratio = ([\d.]+)",
    "平均缓存访问时间": r"average cache access time = ([\d.]+)",
    "平均缓存惩罚时间": r"average cache miss penalty = ([\d.]+)",
    "AMAT": r"cache AMAT = ([\d.]+)"
}


def parse_log(file_path):
    results = {}
    with open(file_path, "r") as file:
        log_data = file.read()

        # 对每个模式应用正则表达式并提取结果
        for key, pattern in patterns.items():
            match = re.search(pattern, log_data)
            if match:
                # 去掉数字中的逗号分隔符，并存储在结果字典中
                results[key] = match.group(1).replace(",", "")
            else:
                results[key] = "N/A"  # 如果没有找到匹配项，标记为"N/A"

    return results


def get_latest_commit_hash():
    # 获取最近一次提交的hash值
    result = subprocess.run(
        ["git", "rev-parse", "HEAD"], capture_output=True, text=True
    )
    return result.stdout.strip()


def update_excel(file_path, commit_hash, description, data):
    # 打开Excel文件
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 找到最后一行
        next_row = sheet.max_row + 1

        # 添加数据
        sheet.append(
            [commit_hash, description, None, None] + [data[key] for key in patterns.keys()]
        )

        # 获取上一行的格式
        if next_row > 2:  # 确保有上一行可以参考
            for col in range(1, sheet.max_column + 1):
                previous_cell = sheet.cell(row=next_row - 1, column=col)
                current_cell = sheet.cell(row=next_row, column=col)
                current_cell.style = copy(previous_cell.style)

        # 保存Excel文件
        workbook.save(file_path)
        print(f"数据已成功添加到 {file_path} 的第 {next_row} 行")
    except Exception as e:
        print(f"无法更新 Excel 文件: {e}")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("用法: python statisticians.py <说明> <log文件路径> <xlsx文件路径>")
        sys.exit(1)

    description = sys.argv[1]
    log_file = sys.argv[2]
    excel_file = sys.argv[3]

    # 解析日志文件
    extracted_data = parse_log(log_file)

    # 获取提交哈希
    commit_hash = get_latest_commit_hash()

    # 更新Excel文件
    update_excel(excel_file, commit_hash, description, extracted_data)
